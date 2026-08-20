# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
xlsx = r"D:\20.git_research\01.topics\21.shleeclay.github.io\cv\application_info.xlsx"
wb = openpyxl.load_workbook(xlsx)
thin = Side(style="thin", color="D9D9D9"); Bd = Border(left=thin,right=thin,top=thin,bottom=thin)
def style_last(ws):
    for c in ws[ws.max_row]:
        c.font = Font(size=10); c.alignment = Alignment(vertical="top", wrap_text=True); c.border = Bd

TITLE = "Bi-temporal ALS assessment of vertical–horizontal forest structures and their structural association in a temperate urban forest"

# ---- Publications 추가 (No.10) ----
ws = wb["Publications"]; h=[c.value for c in ws[1]]
col={n:h.index(n) for n in h}
if any(r[col['DOI']].value=="10.1080/21580103.2026.2712965" for r in ws.iter_rows(min_row=2)):
    print("이미 Publications에 존재 — 스킵")
else:
    row=[None]*len(h)
    row[col['No.']]=10; row[col['Year']]=2026; row[col['Role']]="first"; row[col['Index']]="SCI"
    row[col['IF']]="2.4 (2026)"; row[col['Quartile']]="Q2 (2026)"
    row[col['Journal(EN)']]="Forest Science and Technology"
    row[col['DOI']]="10.1080/21580103.2026.2712965"
    row[col['Authors(EN)']]="Seunghyeon Lee, Yonghwan Kim, Dohee Kim, Hansoo Kim, Youngkeun Song"
    row[col['교신저자(KO)']]="송영근"; row[col['교신저자(EN)']]="Youngkeun Song"
    row[col['논문제목(KO)']]=TITLE; row[col['Title(EN)']]=TITLE
    row[col['게재일자']]="2026. 08. 17"; row[col['저자순위(본인)']]=1; row[col['총저자수']]=5
    ws.append(row); style_last(ws)
    print("Publications No.10 추가")

# ---- Under Review 제거 + 재번호 ----
ws = wb["Under Review"]; h=[c.value for c in ws[1]]; jc=h.index('Journal(EN)'); nc=h.index('No.')
rows=[[c.value for c in r] for r in ws.iter_rows(min_row=2)]
kept=[rr for rr in rows if 'Forest Science and Technology' not in str(rr[jc])]
removed=len(rows)-len(kept)
# 시트 데이터 재작성
ws.delete_rows(2, ws.max_row-1)
for i,rr in enumerate(kept,1):
    rr[nc]=i; ws.append(rr); style_last(ws)
print(f"Under Review 제거 {removed}건, 재번호 → {len(kept)}건")

wb.save(xlsx)
print("xlsx 저장 완료")
