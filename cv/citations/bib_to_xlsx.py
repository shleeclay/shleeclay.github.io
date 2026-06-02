# -*- coding: utf-8 -*-
"""
bib_to_xlsx.py — publications.bib 를 엑셀(publications.xlsx)로 내보냄.
빠른 확인용(DOI 클릭 링크 포함). merge 후 실행:
  python merge_bib.py && python bib_to_xlsx.py
"""
import os, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import merge_bib as M

HERE = os.path.dirname(os.path.abspath(__file__))
SRC  = os.path.join(HERE, "publications.bib")
OUT  = os.path.join(HERE, "publications.xlsx")

MN = ["", "Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

def clean(s):
    """Strip common LaTeX/BibTeX artifacts for display."""
    return (s.replace("\\&", "&").replace("--", "–").replace("{", "").replace("}", "").strip())

rows = []
for raw in M.split_entries(open(SRC, encoding="utf-8").read()):
    if not raw.lstrip().startswith("@"):
        continue
    yr  = M.field(raw, "year")
    mo  = M.month_num(raw)
    doi = M.get_doi(raw)
    au  = M.field(raw, "author")
    fa = au.split(" and ")[0].lower()          # handles "Lee, Seunghyeon" and "Seunghyeon Lee"
    lee_first = ("lee" in fa and "seunghyeon" in fa)
    rows.append({
        "year": int(yr) if yr.isdigit() else 0,
        "month": MN[mo] if 0 < mo < 13 else "",
        "journal": clean(M.field(raw, "journal")),
        "title": clean(M.field(raw, "title")),
        "authors": clean(au),
        "lee_first": "●" if lee_first else "",
        "vol": M.field(raw, "volume"),
        "issue": M.field(raw, "number"),
        "pages": M.field(raw, "pages"),
        "doi": doi,
    })
rows.sort(key=lambda r: (r["year"], 0))  # already ascending in file; keep stable

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "publications"
cols = [("No", 5), ("Year", 7), ("Mon", 6), ("1저자", 7), ("Journal", 34),
        ("Title", 70), ("Authors", 46), ("Vol", 6), ("Issue", 7), ("Pages", 9), ("DOI", 30)]
head_fill = PatternFill("solid", fgColor="1F4E3D")
head_font = Font(bold=True, color="FFFFFF", size=10)
thin = Side(style="thin", color="DDDDDD"); border = Border(*(thin,)*4)

for c, (name, w) in enumerate(cols, 1):
    cell = ws.cell(1, c, name)
    cell.fill = head_fill; cell.font = head_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

for i, r in enumerate(rows, 1):
    vals = [i, r["year"], r["month"], r["lee_first"], r["journal"], r["title"],
            r["authors"], r["vol"], r["issue"], r["pages"], r["doi"]]
    rownum = i + 1
    for c, v in enumerate(vals, 1):
        cell = ws.cell(rownum, c, v)
        cell.font = Font(size=9)
        cell.border = border
        cell.alignment = Alignment(vertical="top",
            wrap_text=(c in (6, 7)),
            horizontal=("center" if c in (1, 2, 3, 4, 8, 9, 10) else "left"))
    doi = r["doi"]
    if doi:
        dcell = ws.cell(rownum, 11)
        dcell.hyperlink = "https://doi.org/" + doi
        dcell.font = Font(size=9, color="1155CC", underline="single")

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:K{len(rows)+1}"
wb.save(OUT)
print(f"wrote {OUT}  ({len(rows)} publications)")
