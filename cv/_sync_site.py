# -*- coding: utf-8 -*-
"""application_info.xlsx 의 사실 데이터를 ../src/data/site.json 에 반영한다.

  설계 원칙
   1) 병합이지 생성이 아니다. logo / cover / pdf / url / id 같은 표시용 필드는 건드리지 않는다.
   2) 빈 값으로 덮지 않는다. xlsx 칸이 비어 있으면 site.json 값을 그대로 둔다.
   3) 항목을 추가·삭제하지 않는다. 기존 항목의 필드만 갱신한다 (추가는 사람이 판단).
   4) 기본은 미리보기다. --write 를 줘야 파일을 쓴다.

  실행:  python _sync_site.py           변경될 내용만 출력
         python _sync_site.py --write   실제 반영 (쓰기 전 .bak 생성)
"""
import io
import json
import os
import re
import shutil
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "application_info.xlsx")
SITE = os.path.join(HERE, "..", "src", "data", "site.json")

wb = openpyxl.load_workbook(XLSX, data_only=True)
site = json.load(open(SITE, encoding="utf-8"))
CHANGES = []
UNMATCHED = []


def rows(name, where=None):
    ws = wb[name]
    hdr = [str(c.value) for c in ws[1]]
    out = [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
    out = [r for r in out if any(v is not None for v in r.values())]
    return [r for r in out if where(r)] if where else out


def norm(v):
    return re.sub(r"[^a-z0-9]", "", str(v or "").lower())


def blank(v):
    return v is None or str(v).strip() == ""


def put(obj, key, val, where):
    """빈 값으로 덮지 않는다. 값이 같으면 아무것도 하지 않는다."""
    if blank(val):
        return
    old = obj.get(key)
    if isinstance(old, dict) or isinstance(old, list):
        return                                   # ko/en 쌍·배열은 아래 전용 처리
    if str(old or "") == str(val):
        return
    CHANGES.append((where, key, old, val))
    obj[key] = val


def put_pair(obj, key, ko, en, where):
    """{'ko':..,'en':..} 형태 필드. 한쪽만 있어도 그쪽만 갱신한다."""
    cur = obj.get(key)
    if not isinstance(cur, dict):
        return
    for sub, val in (("ko", ko), ("en", en)):
        if blank(val) or str(cur.get(sub) or "") == str(val):
            continue
        CHANGES.append((where, f"{key}.{sub}", cur.get(sub), val))
        cur[sub] = val


def put_list(obj, key, val, where):
    """콤마 구분 문자열 -> 배열 필드. ko/en 쌍이면 손대지 않는다 (put_pair_list 용)."""
    if blank(val) or isinstance(obj.get(key), dict):
        return
    new = [x.strip() for x in str(val).split(",") if x.strip()]
    if obj.get(key) == new:
        return
    CHANGES.append((where, key, obj.get(key), new))
    obj[key] = new


def put_pair_list(obj, key, ko, en, where):
    """{'ko':[..], 'en':[..]} 형태. 한쪽만 갱신해도 다른 쪽을 지우지 않는다."""
    cur = obj.get(key)
    if not isinstance(cur, dict):
        return
    for sub, val in (("ko", ko), ("en", en)):
        if blank(val):
            continue
        new = [x.strip() for x in str(val).split(",") if x.strip()]
        if cur.get(sub) == new:
            continue
        CHANGES.append((where, f"{key}.{sub}", cur.get(sub), new))
        cur[sub] = new


# ── 논문: DOI 로 매칭 ─────────────────────────────────────────────────
xp = {norm(r["DOI"]): r for r in rows("Publications") if r["DOI"]}
for it in site["publications"]["items"]:
    r = xp.get(norm(it.get("doi")))
    if not r:
        continue
    w = f'publications[{it.get("id")}]'
    put(it, "year", r["Year"], w)
    put(it, "role", r["Role"], w)
    put(it, "if", r["IF"], w)
    put(it, "quartile", r["Quartile"], w)
    put(it, "journal", r["Journal(EN)"], w)
    put(it, "journalKo", r["학술지(KO)"], w)
    put(it, "date", str(r["게재일자"] or "").replace(". ", "/").replace(".", ""), w)
    put_list(it, "authors", r["Authors(EN)"], w)
    put_list(it, "authorsKo", r["저자(KO)"], w)
    put_pair(it, "title", r["논문제목(KO)"], r["Title(EN)"], w)

# ── 학회: 날짜 + 제목 앞부분으로 매칭 ────────────────────────────────
def ckey(d, t):
    return f'{str(d or "")[:10]}|{norm(t)[:24]}'


_conf = rows("Conferences")
xc = {ckey(r["Date"], r["Title(EN)"]): r for r in _conf}
xc_date = {}
for r in _conf:                       # 같은 날짜에 2건 이상이면 날짜 재시도를 막는다
    xc_date.setdefault(str(r["Date"] or "")[:10], []).append(r)
for it in site["conferences"]["items"]:
    r = xc.get(ckey(it.get("date"), (it.get("title") or {}).get("en")))
    if not r:
        same = xc_date.get(str(it.get("date") or "")[:10], [])
        r = same[0] if len(same) == 1 else None
        if r:
            UNMATCHED.append(f'conferences[{it.get("id")}] 제목이 달라 날짜로 매칭: '
                             f'{(it.get("title") or {}).get("en","")[:56]}')
    if not r:
        UNMATCHED.append(f'conferences[{it.get("id")}] 매칭 실패 — 건너뜀')
        continue
    w = f'conferences[{it.get("id")}]'
    put(it, "year", r["Year"], w)
    put(it, "type", r["Type"], w)
    put(it, "scope", r["Scope"], w)
    put_pair(it, "conference", r["학회(KO)"], r["Conference(EN)"], w)
    put_pair(it, "venue", r["장소(KO)"], r["Venue(EN)"], w)
    put_pair(it, "title", r["발표제목(KO)"], r["Title(EN)"], w)
    put_list(it, "authors", r["Authors"], w)

# ── 자격증: 취득일로 매칭 ────────────────────────────────────────────
xcert = {str(r["취득일"] or "")[:10]: r for r in rows("Certifications")}
for it in site["honors"]["certifications"]["items"]:
    r = xcert.get(str(it.get("date") or "")[:10])
    if r:
        put_pair(it, "name", r["자격증명(KO)"], r["Certification(EN)"],
                 f'certifications[{it.get("date")}]')

# ── 장학: 재단명으로 매칭 ────────────────────────────────────────────
xf = {norm(r["Foundation(EN)"]): r for r in rows("Funding")}
for it in site["honors"]["scholarships"]["items"]:
    r = xf.get(norm((it.get("foundation") or {}).get("en")))
    if not r:
        continue
    w = f'scholarships[{(it.get("foundation") or {}).get("en")}]'
    put(it, "period", r["Period"], w)
    put(it, "fundingUsd", r["Amount(USD)"], w)
    put_pair(it, "name", r["명칭(KO)"], r["Name(EN)"], w)
    put_pair(it, "support", r["지원(KO)"], r["Support(EN)"], w)

# ── 경력: 웹 표시=Y 인 행만, 기관명으로 매칭 ─────────────────────────
xw = {norm(r["Institution Name(EN)"])[:12]: r
      for r in rows("Work & Internships", lambda r: str(r["웹 표시"]).upper() == "Y")}
for it in site["career"]["items"]:
    ck = norm((it.get("company") or {}).get("en"))[:12]
    r = next((v for k, v in xw.items() if k and (k in ck or ck in k)), None)
    if not r:
        continue
    w = f'career[{it.get("id")}]'
    put(it, "fundingUsd", r["지원금(USD)"], w)
    put(it, "papers", r["연계 논문(EN)"], w)
    put(it, "patents", r["연계 특허(EN)"], w)
    put_pair(it, "type", r["고용형태(KO)"], r["고용형태(EN)"], w)

# ── 특허: 출원일로 매칭 ──────────────────────────────────────────────
xpat = {str(r["App Date"] or "")[:10]: r for r in rows("Patents")}
for it in site["patents"]["items"]:
    r = xpat.get(str(it.get("applicationDate") or "")[:10])
    if not r:
        continue
    w = f'patents[{it.get("id")}]'
    put(it, "registrationDate", r["Reg Date"], w)
    put(it, "status", r["Status"], w)
    put_pair(it, "title", r["특허명(KO)"], r["Title(EN)"], w)
    put_pair_list(it, "inventors", r["발명자(KO)"], r["Inventors(EN)"], w)

# ── 보고 ─────────────────────────────────────────────────────────────
if UNMATCHED:
    print("=" * 78)
    print(f"매칭 주의 {len(UNMATCHED)}건")
    print("=" * 78)
    for m in UNMATCHED:
        print("  <!>", m)
    print()
print("=" * 78)
if not CHANGES:
    print("site.json 에 반영할 변경 없음 — xlsx 와 일치합니다.")
    print("=" * 78)
    sys.exit(0)
print(f"site.json 에 반영될 변경 {len(CHANGES)}건")
print("=" * 78)
for where, key, old, new_ in CHANGES:
    print(f"  {where}  {key}")
    print(f"      기존: {str(old)[:96]}")
    print(f"      변경: {str(new_)[:96]}")

if "--write" not in sys.argv:
    print()
    print("미리보기입니다. 반영하려면  python _sync_site.py --write")
    sys.exit(1)

# ── 쓰기: 원문을 국소 치환한다 ───────────────────────────────────────
# json.dump 로 다시 쓰면 손으로 맞춘 빈 줄·그룹핑이 전부 사라져(약 2,900줄)
# 실제 변경을 리뷰할 수 없게 된다. 그래서 문자열 리터럴만 바꾼다.
raw = io.open(SITE, encoding="utf-8").read()
edits, skipped = 0, []


def swap(old, new_, label):
    """따옴표까지 포함한 JSON 문자열 리터럴을 치환. 유일할 때만 바꾼다."""
    global raw, edits
    a_ = json.dumps(str(old), ensure_ascii=False)
    b_ = json.dumps(str(new_), ensure_ascii=False)
    n = raw.count(a_)
    if n == 0:
        skipped.append(f"{label}: 원문에서 찾지 못함")
        return
    if n > 1:
        # 같은 줄(=같은 항목) 안의 중복이면 전부 바꾼다.
        # 국문 제목이 없어 ko/en 에 같은 영문이 들어간 경우가 여기 해당한다.
        lines = [l for l in raw.splitlines() if a_ in l]
        if len(lines) != 1 or lines[0].count(a_) != n:
            skipped.append(f"{label}: 서로 다른 {n}곳에 등장 — 직접 수정 필요")
            return
        raw = raw.replace(a_, b_)
        edits += n
        return
    raw = raw.replace(a_, b_, 1)
    edits += 1


for where, key, old, new_ in CHANGES:
    if isinstance(old, list) and isinstance(new_, list):
        for o, n_ in zip(old, new_):
            if o != n_:
                swap(o, n_, f"{where}.{key}")
    else:
        swap(old, new_, f"{where}.{key}")

shutil.copy2(SITE, SITE + ".bak")
io.open(SITE, "w", encoding="utf-8", newline="").write(raw)
json.load(open(SITE, encoding="utf-8"))          # 파싱 가능한지 즉시 확인
print()
print(f"치환 {edits}건 반영. 백업: {os.path.basename(SITE)}.bak")
for m in skipped:
    print("  <!>", m)
print("git diff src/data/site.json 으로 확인한 뒤 커밋하세요.")
