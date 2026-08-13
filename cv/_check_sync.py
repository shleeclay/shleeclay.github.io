# -*- coding: utf-8 -*-
"""세 곳의 데이터가 어긋났는지 대조해서 보고한다. 아무것도 고치지 않는다.

  대조 대상
    1) cv/application_info.xlsx        — 사실 정본 (18시트)
    2) ../src/data/site.json           — 웹페이지
    3) cv/build_cv.py                  — 영문 CV (+ citations/publications.bib)

  실행:  python _check_sync.py
  종료코드: 0 = 이상 없음, 1 = 확인 필요 항목 있음
"""
import ast
import json
import os
import re
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "application_info.xlsx")
SITE = os.path.join(HERE, "..", "src", "data", "site.json")
CVPY = os.path.join(HERE, "build_cv.py")
BIB = os.path.join(HERE, "citations", "publications.bib")

WARN = []


def warn(msg):
    WARN.append(msg)
    return "  <!>"


# ── 소스 로드 ─────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(XLSX, data_only=True)
site = json.load(open(SITE, encoding="utf-8"))
cv_src = open(CVPY, encoding="utf-8").read()
cv_ast = ast.parse(cv_src)


def cv_list(name):
    """build_cv.py 최상위 리스트 변수를 값으로 가져온다."""
    for n in cv_ast.body:
        if isinstance(n, ast.Assign) and isinstance(n.targets[0], ast.Name) \
                and n.targets[0].id == name and isinstance(n.value, ast.List):
            try:
                return ast.literal_eval(n.value)
            except Exception:
                return None
    return None


def cv_inline(func, nth=0):
    """build_cv.py 안에서 func([...]) 형태로 인라인된 리스트를 가져온다."""
    hits = []
    for n in ast.walk(cv_ast):
        if isinstance(n, ast.Call) and isinstance(n.func, ast.Name) and n.func.id == func \
                and n.args and isinstance(n.args[0], ast.List):
            try:
                hits.append((n.lineno, ast.literal_eval(n.args[0])))
            except Exception:
                pass
    hits.sort()
    return hits[nth][1] if nth < len(hits) else None


def sj(path):
    o = site
    for k in path.split("."):
        o = o[k]
    return o


def rows(sheet):
    """시트가 없으면 None. 시트 삭제·개명도 보고 대상이므로 크래시하지 않는다."""
    if sheet not in wb.sheetnames:
        return None
    return wb[sheet].max_row - 1


def bib_dois():
    txt = open(BIB, encoding="utf-8").read()
    return {m.group(1).lower().strip() for m in re.finditer(r"doi\s*=\s*\{([^}]*)\}", txt)}


def col(sheet, name):
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    hdr = [str(c.value) for c in ws[1]]
    if name not in hdr:
        return []
    j = hdr.index(name)
    return [r[j] for r in ws.iter_rows(min_row=2, values_only=True)]


# ── A. 항목 수 대조 ───────────────────────────────────────────────────
# (xlsx시트, site.json경로, CV값, 의도된 차이 설명 or None)
SPEC = [
    ("Research Interests", "hero.focus", None),
    ("Education", "education.items", "xlsx만 고등학교 행 포함 (CV 표시=N)"),
    ("Publications", "publications.items", None),
    ("Under Review", "publications.underReview.items", None),
    ("Books", "honors.books.items", None),
    ("Conferences", "conferences.items", None),
    ("Invited Talks", None, "학회 시트의 CV분류=invited 와 합산해 CV 4건"),
    ("Teaching", "teaching.items", None),
    ("Funding", "honors.scholarships.items", None),
    ("Projects", "projects.items", None),
    ("Patents", "patents.items", None),
    ("Awards", None, "웹에 미게재"),
    ("Service & Membership", None, "웹에 미게재"),
    ("Certifications", "honors.certifications.items", None),
    ("Language Tests", None, None),
    ("Technical Skills", None, "build_cv_short.py 에만 존재"),
    ("Work & Internships", None, "웹 노출은 아래 플래그 대조 참조"),
    ("Personal", None, "인적사항은 CV/웹 대상 아님"),
]

print("=" * 78)
print("A. 항목 수 대조")
print("=" * 78)
print(f"  {'항목':24s} {'xlsx':>5s} {'web':>5s}   비고")
for sh, jpath, note in SPEC:
    x = rows(sh)
    w = len(sj(jpath)) if jpath else None
    f = lambda v: "-" if v is None else str(v)
    if x is None:
        warn(f"{sh}: xlsx 에 시트가 없음 (삭제되었거나 이름이 바뀜)")
        print(f"  {sh:24s} {'없음':>5s} {f(w):>5s}   <!>")
        continue
    flag = ""
    if w is not None and w != x:
        flag = f"  <{note}>" if note else warn(f"{sh}: xlsx={x} web={w}")
    elif note:
        flag = f"  <{note}>"
    print(f"  {sh:24s} {x:>5d} {f(w):>5s} {flag}")

# ── A-2. build_cv.py 가 xlsx 를 읽는지 (리팩터 회귀 방지) ────────────
print()
print("  build_cv.py 데이터 출처")
_reads = sorted({ast.literal_eval(n.args[0])
                 for n in ast.walk(cv_ast)
                 if isinstance(n, ast.Call) and isinstance(n.func, ast.Name)
                 and n.func.id == "sheet" and n.args})
_lits = [getattr(n.targets[0], "id", "?") for n in cv_ast.body
         if isinstance(n, ast.Assign) and isinstance(n.value, (ast.List, ast.Tuple))
         and not getattr(n.targets[0], "id", "").startswith("_KEEP")]
print(f"    xlsx 를 읽는 시트: {', '.join(_reads)}")
print(f"    남은 리터럴: {', '.join(_lits) if _lits else '없음'}"
      "   (contact/_pubs 는 정상 — 연락처와 bib 누적용)")
_HARDCODED = ["Research Interests", "Books", "Teaching Experience", "References"]
print(f"    아직 코드에 서술된 섹션: {', '.join(_HARDCODED)}")

# ── A-3. 경력 — 웹 표시 플래그 기준 대조 ──────────────────────────────
print()
print("  경력 웹 노출: xlsx 의 '웹 표시'=Y 인 행이 site.json career 와 맞는지")
flags = col("Work & Internships", "웹 표시")
n_y = sum(1 for v in flags if str(v).strip().upper() == "Y")
n_web = len(sj("career.items"))
if n_y == n_web:
    print(f"    Y={n_y} · web={n_web}  일치")
else:
    warn(f"경력 웹 노출: 웹표시Y={n_y} 인데 site.json career={n_web}")
    print(f"    Y={n_y} · web={n_web}  <!>")
blank = [i for i, v in enumerate(flags, 2) if not str(v or "").strip()]
if blank:
    warn(f"경력 '웹 표시' 미지정 행: {blank}")
    print(f"    '웹 표시' 빈 칸 행: {blank}  <!>")

# ── B. 논문 DOI 집합 대조 ─────────────────────────────────────────────
print()
print("=" * 78)
print("B. 논문 DOI 집합 (bib · web · xlsx)")
print("=" * 78)
d_bib = bib_dois()
d_web = {(it.get("doi") or "").lower().strip() for it in sj("publications.items") if it.get("doi")}
d_xls = {str(v).lower().strip() for v in col("Publications", "DOI") if v}
for label, a, b in [("bib에만", d_bib, d_web | d_xls),
                    ("web에만", d_web, d_bib | d_xls),
                    ("xlsx에만", d_xls, d_bib | d_web)]:
    only = a - b
    if only:
        warn(f"논문 DOI {label}: {sorted(only)}")
        print(f"  {label}: {sorted(only)}")
if not (d_bib ^ d_web) and not (d_bib ^ d_xls):
    print(f"  세 곳 모두 {len(d_bib)}편 일치")

# ── C. 값 불일치 (IF · quartile) ──────────────────────────────────────
print()
print("=" * 78)
print("C. 값 대조 — 논문 IF · Quartile (web ↔ xlsx, DOI 기준)")
print("=" * 78)
web_if = {(it.get("doi") or "").lower(): (str(it.get("if") or ""), str(it.get("quartile") or ""))
          for it in sj("publications.items")}
ws = wb["Publications"]
hdr = [str(c.value) for c in ws[1]]
bad = 0
for r in ws.iter_rows(min_row=2, values_only=True):
    doi = str(r[hdr.index("DOI")] or "").lower().strip()
    xi = (str(r[hdr.index("IF")] or ""), str(r[hdr.index("Quartile")] or ""))
    wi = web_if.get(doi)
    if wi and xi != wi:
        bad += 1
        warn(f"IF/Q 불일치 {doi}: xlsx={xi} web={wi}")
        print(f"  {doi}\n      xlsx={xi}   web={wi}")
if not bad:
    print("  불일치 없음")

# ── D. xlsx 빈 칸 (사용자 입력 대기) ──────────────────────────────────
print()
print("=" * 78)
print("D. xlsx 미입력 — 전부 비어 있는 컬럼")
print("=" * 78)
for sh in wb.sheetnames:
    w = wb[sh]
    n = w.max_row - 1
    if n == 0:
        print(f"  {sh:24s} (시트 전체 비어 있음)")
        continue
    empty = []
    for j, h in enumerate([c.value for c in w[1]], 1):
        vals = [w.cell(row=i, column=j).value for i in range(2, w.max_row + 1)]
        if all(v is None or str(v).strip() == "" for v in vals):
            empty.append(str(h))
    if empty:
        print(f"  {sh:24s} {', '.join(empty)}")

# ── 요약 ──────────────────────────────────────────────────────────────
print()
print("=" * 78)
if WARN:
    print(f"확인 필요 {len(WARN)}건")
    for w_ in WARN:
        print("  <!>", w_)
else:
    print("확인 필요 항목 없음")
print("=" * 78)
sys.exit(1 if WARN else 0)
